"""
📊 ENHANCED Excel Results Generator
Generates a formatted Excel file with charts and analysis
Requires: pip install openpyxl
"""

from mcp_client import MCPClient
import re
from datetime import datetime
import csv

# Check if openpyxl is available for Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not installed. Will generate CSV only.")
    print("   To create Excel with charts: pip install openpyxl")

# ============================================================
# YOUR QUESTIONS
# ============================================================

MY_QUESTIONS = [
    "What is the maximum number of analog line-level outputs on the EX-1280, and is this different from the EX-1280C?",
    "Comparing the DM8SE and DM6PE, which one has the lowest Net Weight in kg? If this lighter speaker is installed outside, what is its stated IP rating?",
    "We need the lowest latency possible, and redundancy is critical. Between the EX-1280 and EX-1280C, which processor can offer dual-port Dante redundancy and the best guaranteed audio latency (Analog In to Out)?",
    "If I must wire a speaker in its 100V maximum tap setting, which DM speaker model can handle the least amount of long-term continuous power, and what is that value in Watts?",
    "If a project strictly requires the use of an RS-232 serial control port for system integration, does the model EX-1280 offer this feature, or is it exclusive to the conferencing EX-1280C?",
    "Which of the two DesignMax speakers has the lowest Net Weight in kg? If this lighter speaker is installed outside, what is its stated IP rating?",
    "What is the list price for the ControlSpace EX-1280 when purchased directly from a Bose distributor?",
    "What is the official warranty length for the DM6PE, and does that warranty cover installation damage?",
    "What is the minimum required NovaOS firmware version needed for the EX-1280C to handle its 2-line VoIP feature?",
    "How does the sound dispersion pattern of the DM8SE compare to the DM6PE in a small room, and which one has a more powerful bass response?",
    "What are the planned successor models for the ControlSpace EX-1280 line scheduled for release in Q1 2026?",
    "What is the maximum input level (in dBu) on the EX-1280C, and what is the smallest transformer tap available on the DM6PE for a 70 V system?",
    "If I use the EX-1280 to process audio for four DM8SE speakers wired to four different analog outputs, how many total unused analog inputs will remain on the processor?",
    "What is the Maximum Input Voltage required for the EX-1280C? Can the DM8SE loudspeaker withstand the same voltage in its 100 V tap setting?",
    "Using the EX-1280C, if I configure all analog inputs for +48V phantom power, how many GPO outputs are available to interface with a third-party control system?",
    "We need to use the speaker with the lowest overall physical depth (D) and connect it to the processor that is designed specifically for an office conference room. Name the specific model and state its depth."
]

# ============================================================
# Query all questions
# ============================================================

print("\n" + "="*70)
print("  QUERYING RAG SYSTEM".center(70))
print("="*70 + "\n")

client = MCPClient("http://localhost:8001")
results = []

for i, question in enumerate(MY_QUESTIONS, 1):
    print(f"[{i}/{len(MY_QUESTIONS)}] Processing: {question[:60]}...")
    
    try:
        response = client.query_documentation(question)
        
        if 'content' in response and len(response['content']) > 0:
            full_response = "".join(item.get('text', '') for item in response['content'])
            
            # Extract components
            answer_match = re.search(r'\*\*Answer:\*\*\s*(.*?)(?=\n\n|\*\*)', full_response, re.DOTALL)
            confidence_match = re.search(r'\*\*Confidence:\*\*\s*(\d+)%\s*\((\w+)\)', full_response)
            time_match = re.search(r'\*\*Query Time:\*\*\s*([\d.]+)s', full_response)
            sources_match = re.search(r'\*\*Sources:\*\*\s*(.*?)(?=\n\n|\*\*Query Time)', full_response, re.DOTALL)
            
            answer = answer_match.group(1).strip() if answer_match else "No answer extracted"
            confidence_pct = int(confidence_match.group(1)) if confidence_match else 0
            confidence_label = confidence_match.group(2) if confidence_match else "N/A"
            query_time = float(time_match.group(1)) if time_match else 0.0
            sources = sources_match.group(1).strip() if sources_match else "N/A"
            
            results.append({
                'Question #': i,
                'Question': question,
                'Answer': answer,
                'Confidence (%)': confidence_pct,
                'Confidence Level': confidence_label,
                'Time Taken (s)': query_time,
                'Sources': sources.replace('\n', ' | ')
            })
            
            print(f"  ✓ {confidence_pct}% confidence in {query_time:.2f}s")
        else:
            results.append({
                'Question #': i,
                'Question': question,
                'Answer': 'ERROR: No response',
                'Confidence (%)': 0,
                'Confidence Level': 'N/A',
                'Time Taken (s)': 0.0,
                'Sources': 'N/A'
            })
            print(f"  ✗ No response")
            
    except Exception as e:
        print(f"  ✗ Error: {e}")
        results.append({
            'Question #': i,
            'Question': question,
            'Answer': f'ERROR: {str(e)}',
            'Confidence (%)': 0,
            'Confidence Level': 'N/A',
            'Time Taken (s)': 0.0,
            'Sources': 'N/A'
        })

print("\n" + "="*70)
print("  GENERATING REPORTS".center(70))
print("="*70 + "\n")

# ============================================================
# Generate CSV (always)
# ============================================================

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
csv_filename = f"bose_rag_results_{timestamp}.csv"

print(f"📄 Generating CSV: {csv_filename}")

with open(csv_filename, 'w', newline='', encoding='utf-8') as csvfile:
    fieldnames = ['Question #', 'Question', 'Answer', 'Confidence (%)', 'Confidence Level', 'Time Taken (s)', 'Sources']
    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(results)

print(f"✅ CSV saved: {csv_filename}")

# ============================================================
# Generate Excel with formatting and charts
# ============================================================

if EXCEL_AVAILABLE and results:
    excel_filename = f"bose_rag_results_{timestamp}.xlsx"
    print(f"\n📊 Generating Excel: {excel_filename}")
    
    wb = Workbook()
    
    # ===== Sheet 1: Results =====
    ws_results = wb.active
    ws_results.title = "Results"
    
    # Headers
    headers = ['Q#', 'Question', 'Answer', 'Confidence %', 'Level', 'Time (s)', 'Sources']
    ws_results.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws_results[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data with conditional formatting
    for row in results:
        ws_results.append([
            row['Question #'],
            row['Question'],
            row['Answer'],
            row['Confidence (%)'],
            row['Confidence Level'],
            row['Time Taken (s)'],
            row['Sources']
        ])
    
    # Apply conditional formatting for confidence
    for row_idx in range(2, len(results) + 2):
        conf_cell = ws_results.cell(row=row_idx, column=4)
        level_cell = ws_results.cell(row=row_idx, column=5)
        
        if isinstance(conf_cell.value, int):
            if conf_cell.value >= 85:
                level_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                level_cell.font = Font(color="006100", bold=True)
            elif conf_cell.value >= 70:
                level_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
                level_cell.font = Font(color="9C6500", bold=True)
            else:
                level_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
                level_cell.font = Font(color="9C0006", bold=True)
    
    # Adjust column widths
    ws_results.column_dimensions['A'].width = 5
    ws_results.column_dimensions['B'].width = 60
    ws_results.column_dimensions['C'].width = 80
    ws_results.column_dimensions['D'].width = 12
    ws_results.column_dimensions['E'].width = 12
    ws_results.column_dimensions['F'].width = 12
    ws_results.column_dimensions['G'].width = 40
    
    # ===== Sheet 2: Summary Statistics =====
    ws_summary = wb.create_sheet("Summary")
    
    # Calculate stats
    valid_confidences = [r['Confidence (%)'] for r in results if isinstance(r['Confidence (%)'], int) and r['Confidence (%)'] > 0]
    valid_times = [r['Time Taken (s)'] for r in results if isinstance(r['Time Taken (s)'], (int, float)) and r['Time Taken (s)'] > 0]
    
    avg_confidence = sum(valid_confidences) / len(valid_confidences) if valid_confidences else 0
    avg_time = sum(valid_times) / len(valid_times) if valid_times else 0
    
    high_conf = sum(1 for r in results if r['Confidence Level'] == 'high')
    medium_conf = sum(1 for r in results if r['Confidence Level'] == 'medium')
    low_conf = sum(1 for r in results if r['Confidence Level'] in ['low', 'very_low'])
    
    # Add summary data
    summary_data = [
        ["Bose RAG System - Performance Summary", ""],
        ["", ""],
        ["Metric", "Value"],
        ["Total Questions", len(results)],
        ["Average Confidence", f"{avg_confidence:.1f}%"],
        ["Average Query Time", f"{avg_time:.2f}s"],
        ["", ""],
        ["Confidence Distribution", "Count"],
        ["High Confidence (≥85%)", high_conf],
        ["Medium Confidence (70-84%)", medium_conf],
        ["Low Confidence (<70%)", low_conf],
    ]
    
    for row in summary_data:
        ws_summary.append(row)
    
    # Style summary
    ws_summary['A1'].font = Font(bold=True, size=16, color="4472C4")
    ws_summary.merge_cells('A1:B1')
    
    for row in [3, 8]:
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        ws_summary[f'B{row}'].font = Font(bold=True, size=12)
        ws_summary[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws_summary[f'B{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 15
    
    # Add pie chart for confidence distribution
    pie = PieChart()
    labels = Reference(ws_summary, min_col=1, min_row=9, max_row=11)
    data = Reference(ws_summary, min_col=2, min_row=8, max_row=11)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Confidence Distribution"
    ws_summary.add_chart(pie, "D3")
    
    wb.save(excel_filename)
    print(f"✅ Excel saved: {excel_filename}")
    print(f"\n📊 Summary Statistics:")
    print(f"   Total Questions: {len(results)}")
    print(f"   Average Confidence: {avg_confidence:.1f}%")
    print(f"   Average Query Time: {avg_time:.2f}s")
    print(f"   High Confidence: {high_conf} ({high_conf/len(results)*100:.1f}%)")
    print(f"   Medium Confidence: {medium_conf} ({medium_conf/len(results)*100:.1f}%)")
    print(f"   Low Confidence: {low_conf} ({low_conf/len(results)*100:.1f}%)")

print("\n" + "="*70)
print("  COMPLETE!".center(70))
print("="*70)
print("\n💡 Files generated:")
print(f"   📄 {csv_filename} (CSV - open in any spreadsheet)")
if EXCEL_AVAILABLE:
    print(f"   📊 {excel_filename} (Excel with charts and formatting)")
print("\n🎯 Use these files to showcase your RAG system's performance!")
print("   - Add to your presentation")
print("   - Submit with your report")
print("   - Demonstrate reliability with confidence scores")
